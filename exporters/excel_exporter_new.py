import pandas as pd
from typing import List, Dict, Any
from pathlib import Path
import os
from models.game_data import GameData


class ExcelExporter:
    """Exportiert Bundesliga-Daten in Excel-Dateien."""

    def __init__(self, output_dir: str = "exports"):
        """
        Initialisiert den ExcelExporter.

        Args:
            output_dir: Pfad zum Ausgabeverzeichnis (Standard: 'exports')
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def set_output_directory(self, output_dir: str):
        """
        Setzt ein neues Ausgabeverzeichnis.

        Args:
            output_dir: Neuer Pfad zum Ausgabeverzeichnis
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def export_by_team(self, games: List[GameData], filename: str = None) -> str:
        """Exportiert Spiele gruppiert nach Teams in separate Sheets."""
        if not games:
            raise ValueError("Keine Spiele zum Exportieren vorhanden")

        if filename is None:
            filename = f"bundesliga_daten_{len(games)}_spiele.xlsx"

        filepath = self.output_dir / filename

        # Spiele nach Teams gruppieren
        teams_games = {}
        for game in games:
            home_team = game.home_team.name
            away_team = game.away_team.name

            # Spiel dem Heimteam zuordnen
            if home_team not in teams_games:
                teams_games[home_team] = []
            teams_games[home_team].append(game)

            # Spiel dem Auswärtsteam zuordnen
            if away_team not in teams_games:
                teams_games[away_team] = []
            teams_games[away_team].append(game)

        # Excel-Writer erstellen
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Übersichts-Sheet
            self._create_overview_sheet(games, writer)

            # Team-spezifische Sheets
            for team_name, team_games in teams_games.items():
                safe_team_name = self._sanitize_sheet_name(team_name)
                self._create_team_sheet(team_games, team_name, writer, safe_team_name)

            # Statistik-Sheet
            self._create_statistics_sheet(games, writer)

        return str(filepath)

    def _create_overview_sheet(self, games: List[GameData], writer):
        """Erstellt ein Übersichts-Sheet mit allen Spielen."""
        overview_data = []

        for game in games:
            # Verwende die korrekt extrahierte matchday aus dem GameData-Objekt
            row_data = {
                "Datum": game.date,
                "Saison": game.season,
                "Spieltag": game.matchday
                or "",  # KORREKTUR: Verwende game.matchday statt _extract_spieltag
                "Heimteam": game.home_team.name,
                "Auswärtsteam": game.away_team.name,
                "Ergebnis": f"{game.home_score}:{game.away_score}",
                "Tore_Heim": game.home_score,
                "Tore_Auswärts": game.away_score,
                "Tore_Gesamt": game.home_score + game.away_score,
                "Gewinner": self._determine_winner(game),
                "Torschützen_Heim": self._format_goals(game.home_goals),
                "Torschützen_Auswärts": self._format_goals(game.away_goals),
                "Aufstellung_Heim": self._format_lineup(game.home_team),
                "Aufstellung_Auswärts": self._format_lineup(game.away_team),
            }
            overview_data.append(row_data)

        df = pd.DataFrame(overview_data)
        # Sortieren nach Datum
        df = df.sort_values("Datum", ascending=False)

        df.to_excel(writer, sheet_name="Übersicht", index=False)

        # Formatierung
        worksheet = writer.sheets["Übersicht"]
        self._format_worksheet(worksheet, df)

    def _create_team_sheet(
        self, games: List[GameData], team_name: str, writer, sheet_name: str
    ):
        """Erstellt ein Sheet für ein spezifisches Team."""
        team_data = []

        for game in games:
            # Basis-Daten mit korrekter Spieltag-Nummer
            row_data = {
                "Datum": game.date,
                "Saison": game.season,
                "Spieltag": game.matchday
                or "",  # KORREKTUR: Verwende game.matchday statt _extract_spieltag
                "Ergebnis": f"{game.home_score}:{game.away_score}",
                "Tore_Gesamt": game.home_score + game.away_score,
                "Gewinner": self._determine_winner(game),
            }

            # Team-spezifische Torschützen
            if game.home_team.name == team_name:
                # Heimspiel für dieses Team
                row_data["Torschützen_Team"] = self._format_goals(game.home_goals)
                row_data["Torschützen_Gegner"] = self._format_goals(game.away_goals)
                row_data["Aufstellung_Team"] = self._format_lineup(game.home_team)
                row_data["Aufstellung_Gegner"] = self._format_lineup(game.away_team)
            else:
                # Auswärtsspiel für dieses Team
                row_data["Torschützen_Team"] = self._format_goals(game.away_goals)
                row_data["Torschützen_Gegner"] = self._format_goals(game.home_goals)
                row_data["Aufstellung_Team"] = self._format_lineup(game.away_team)
                row_data["Aufstellung_Gegner"] = self._format_lineup(game.home_team)

            team_data.append(row_data)

        df = pd.DataFrame(team_data)
        df = df.sort_values("Datum", ascending=False)

        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Formatierung
        worksheet = writer.sheets[sheet_name]
        self._format_worksheet(worksheet, df)

    def _create_statistics_sheet(self, games: List[GameData], writer):
        """Erstellt ein Sheet mit Spielstatistiken."""
        # Basis-Statistiken berechnen
        total_games = len(games)
        total_goals = sum(game.home_score + game.away_score for game in games)
        avg_goals = total_goals / total_games if total_games > 0 else 0

        # Statistiken in DataFrame umwandeln
        stats_data = [
            {"Statistik": "Gesamtanzahl Spiele", "Wert": total_games},
            {"Statistik": "Gesamtanzahl Tore", "Wert": total_goals},
            {
                "Statistik": "Durchschnittliche Tore pro Spiel",
                "Wert": f"{avg_goals:.2f}",
            },
        ]

        df = pd.DataFrame(stats_data)

        df.to_excel(writer, sheet_name="Statistiken", index=False)

        # Formatierung
        worksheet = writer.sheets["Statistiken"]
        self._format_worksheet(worksheet, df)

    def _sanitize_sheet_name(self, name: str) -> str:
        """Bereinigt Teamnamen für Excel-Sheet-Namen."""
        # Excel-Zeichen entfernen die nicht erlaubt sind
        invalid_chars = ["/", "\\", "?", "*", "[", "]", ":"]
        for char in invalid_chars:
            name = name.replace(char, "_")

        # Länge auf 31 Zeichen begrenzen (Excel-Limit)
        return name[:31]

    def _determine_winner(self, game: GameData) -> str:
        """Bestimmt den Gewinner eines Spiels."""
        if game.home_score > game.away_score:
            return game.home_team.name
        elif game.away_score > game.home_score:
            return game.away_team.name
        else:
            return "Unentschieden"

    def _format_worksheet(self, worksheet, df):
        """Formatiert ein Excel-Arbeitsblatt."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

            # Header formatieren
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Spaltenbreiten anpassen
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Datenzeilen formatieren (abwechselnde Farben)
            light_fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )

            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row_num % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill

        except ImportError:
            # Fallback falls openpyxl-Styles nicht verfügbar sind
            pass

    def _extract_spieltag(self, date_str: str) -> str:
        """
        VERALTETE METHODE - Wird nicht mehr verwendet!
        Die Spieltag-Nummer wird jetzt direkt aus dem HTML extrahiert und in game.matchday gespeichert.
        """
        # Diese Methode ist veraltet, wird aber für Kompatibilität beibehalten
        return ""

    def _format_goals(self, goals: List) -> str:
        """Formatiert eine Liste von Toren als String."""
        if not goals:
            return ""

        return ", ".join([f"{goal.scorer} ({goal.minute}')" for goal in goals])

    def _format_lineup(self, team) -> str:
        """Formatiert die Aufstellung eines Teams als String."""
        if not hasattr(team, "players") or not team.players:
            return ""

        return ", ".join([f"{player.name}" for player in team.players])
